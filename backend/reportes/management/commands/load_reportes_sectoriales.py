import re
from pathlib import Path

from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reportes.models import ImportadorProbable, ReporteSectorial, ReporteSectorialDetalle


def normalized(value):
    return str(value or "").strip().upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace(".", "")


def text(value):
    return "" if value is None else str(value).strip()


class Command(BaseCommand):
    help = "Carga hojas base de reportes sectoriales Excel e importa la referencia de probable importador."

    def add_arguments(self, parser):
        parser.add_argument("paths", nargs="+", help="Archivos XLSX de reportes sectoriales")

    def handle(self, *args, **options):
        for raw_path in options["paths"]:
            path = Path(raw_path)
            if not path.exists():
                raise CommandError(f"No existe: {path}")
            self.load_file(path)
        call_command("rebuild_perfiles_importadores")

    def load_file(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            if not (normalized(sheet.title).startswith("BASE ") or normalized(sheet.title) == "BASE DE DATOS"):
                continue
            self.load_sheet(path, sheet)

    def load_sheet(self, path, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        title = " ".join(text(cell) for cell in rows[0] if cell)
        rubro_line = " ".join(text(cell) for cell in rows[1] if cell)
        period_line = " ".join(text(cell) for cell in rows[2] if cell)
        rubro = rubro_line.split(":", 1)[-1].strip() or sheet.title
        normalized_period = normalized(period_line)
        match = re.search(r"(20\d{2})", normalized_period)
        months = {"ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12}
        if not match:
            raise CommandError(f"No se pudo detectar período en {path.name} / {sheet.title}")
        month_names = re.findall("|".join(months), normalized_period)
        if not month_names:
            raise CommandError(f"No se pudo detectar mes en {path.name} / {sheet.title}")
        periodo_mes, periodo_anio = months[month_names[-1]], int(match.group(1))

        header_index = next((index for index, row in enumerate(rows) if "PROBABLE IMPORTADOR" in [normalized(cell) for cell in row]), None)
        if header_index is None:
            raise CommandError(f"No se encontró encabezado en {path.name} / {sheet.title}")
        headers = {normalized(value): index for index, value in enumerate(rows[header_index]) if value}
        reporte, _ = ReporteSectorial.objects.update_or_create(
            nombre_archivo=path.name,
            hoja_base=sheet.title,
            defaults={"rubro": rubro, "periodo_anio": periodo_anio, "periodo_mes": periodo_mes, "es_acumulado": "ACUMUL" in normalized(sheet.title), "ruta_origen": str(path)},
        )
        ReporteSectorialDetalle.objects.filter(reporte=reporte).delete()
        details = []
        for line_no, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(value not in (None, "") for value in row):
                continue
            value = lambda *names: text(next((row[headers[name]] for name in names if name in headers and headers[name] < len(row)), ""))
            rut = value("RUT", "RUT ")
            dv = value("DV")
            importer_name = value("PROBABLE IMPORTADOR")
            importer = None
            if importer_name:
                importer, _ = ImportadorProbable.objects.get_or_create(rut=rut, dv=dv, nombre=importer_name)
            payload = {text(rows[header_index][index]): text(cell) for index, cell in enumerate(row) if index < len(rows[header_index]) and rows[header_index][index]}
            details.append(ReporteSectorialDetalle(
                reporte=reporte,
                nro_linea=line_no,
                importador_probable=importer,
                rut=rut,
                dv=dv,
                aduana_nombre=value("ADUANA"),
                pais_origen_nombre=value("PAIS DE ORIGEN"),
                partida_arancelaria_codigo=value("ARANCEL"),
                mercaderia=value("MERCADERIA"),
                valor_fob=value("US $ FOB"),
                valor_flete=value("US$ FLETE"),
                valor_seguro=value("US $ SEGURO"),
                valor_cif=value("US$ CIF", "US$ CIF "),
                payload_json=payload,
            ))
        ReporteSectorialDetalle.objects.bulk_create(details, batch_size=1000)
        reporte.total_registros = len(details)
        reporte.save(update_fields=["total_registros"])
        self.stdout.write(self.style.SUCCESS(f"{path.name} / {sheet.title}: {len(details)} registros"))
