from pathlib import Path

from openpyxl import load_workbook
from django.core.management.base import BaseCommand

from catalogos.models import CatalogoCodigo, PartidaArancelaria


class Command(BaseCommand):
    help = "Carga catálogos base desde los xlsx de doc-req"

    def add_arguments(self, parser):
        parser.add_argument("--doc-req", dest="doc_req", default="/datos/proyectos/cec-comex-platform/doc-req")
        parser.add_argument("--only-codes", action="store_true")

    def handle(self, *args, **options):
        base = Path(options["doc_req"])
        tablas = base / "tablas_de_codigos(1).xlsx"
        partidas = base / "Tabla de Partidas Arancelarias(1).xlsx"

        self.load_codigo_catalogs(tablas)
        if not options["only_codes"]:
            self.load_partidas(partidas)
        self.stdout.write(self.style.SUCCESS("Catálogos cargados"))

    def load_codigo_catalogs(self, path: Path):
        wb = load_workbook(path, data_only=True)
        sheet_mapping = {
            "Aduanas": "aduanas",
            "Banco Comercial": "bancos_comerciales",
            "Cláusulas de Compra Venta": "clausulas_compra_venta",
            "Artículos de denuncia": "articulos_denuncia",
            "Comunas": "comunas",
            "Formas de pago": "formas_pago",
            "Forma de pago gravámen": "formas_pago_gravamen",
            "Modalidades de Venta": "modalidades_venta",
            "Monedas": "monedas",
            "Países": "paises",
            "Puertos": "puertos",
            "Regiones": "regiones",
            "Tipo de Bultos": "tipos_bulto",
            "Tipo de cuentas": "tipos_cuenta",
            "Tipos de carga": "tipos_carga",
            "Tipos de operacion Din": "tipos_operacion_din",
            "Unidads de Medida": "unidades_medida",
            "Vía de transporte": "via_transporte",
            "Origen de divisas": "origen_divisas",
            "Vistos Buenos": "vistos_buenos",
            "Régimen de Importación": "regimen_importacion",
            "Claves económicas": "claves_economicas",
            "Zonas económicas": "zonas_economicas",
            "Claves económicas exportación": "claves_economicas_exportacion",
        }
        for sheet_name, grupo in sheet_mapping.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            records = {}
            for row in ws.iter_rows(values_only=True):
                values = [v for v in row if v is not None]
                if len(values) < 2:
                    continue
                codigo = str(values[0]).strip()[:128]
                glosa = str(values[1]).strip()
                if not codigo or codigo.lower().replace("\xa0", "") in {"codigo", "código"}:
                    continue
                records[codigo] = glosa
            existing = {item.codigo: item for item in CatalogoCodigo.objects.filter(grupo=grupo, codigo__in=records)}
            created = []
            updated = []
            for codigo, glosa in records.items():
                item = existing.get(codigo)
                if item:
                    item.glosa = glosa
                    item.vigente = True
                    item.origen = "DOC-REQ"
                    updated.append(item)
                else:
                    created.append(CatalogoCodigo(grupo=grupo, codigo=codigo, glosa=glosa, origen="DOC-REQ"))
            CatalogoCodigo.objects.bulk_create(created, batch_size=1000)
            CatalogoCodigo.objects.bulk_update(updated, ["glosa", "vigente", "origen"], batch_size=1000)
            self.stdout.write(f"{grupo}: {len(created)} creados, {len(updated)} actualizados")

    def load_partidas(self, path: Path):
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = [v for v in row if v is not None]
            if len(values) < 2:
                continue
            codigo = str(values[0]).strip()[:32]
            glosa = str(values[1]).strip() if len(values) > 1 else ""
            if not codigo or codigo.lower() == "codigo" or glosa.lower() == "glosa":
                continue
            rows.append(PartidaArancelaria(codigo=codigo, glosa=glosa, origen="DOC-REQ"))
        if rows:
            PartidaArancelaria.objects.all().delete()
            PartidaArancelaria.objects.bulk_create(rows, batch_size=1000)
